import h5py
import openpyxl


def make_workbook(excel_name):  # This function makes a new workbook and puts the below names into their respective cells
    wb = openpyxl.Workbook()
    ws = wb.active

    ws['A1'] = 'Worm_Index'  # puts 'Worm_Index' on to cell A1. The below lines work equivalently.
    ws['B1'] = 'n_frames'
    ws['C1'] = 'Backwards Time'
    ws['D1'] = 'Backwards Distance'
    ws['E1'] = 'Paused Time'
    ws['F1'] = 'Forward Time'
    ws['G1'] = 'Forward Distance'

    wb.save(str(excel_name))  # Saves the excel file once the column headers have been added.


def print_to_excel(excel_file, dct, i):  # This function will open up an excel file and print each value of dct (the actual info) into it column-wise.
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    v = list(dct.values())

    ws['A{}'.format(i)] = str(v[0])  # This is the part that puts the info for the current worm into the current column. 'A{}' is A2, A3, A4, etc depending on which worm's info is currently being added.
    ws['B{}'.format(i)] = str(v[1])  # Works equivalently for the above, but with cell B1, B2, etc.
    ws['C{}'.format(i)] = str(v[2])  # If you want to make this script fill out a new cell, copy/paste one of these lines and adjust accordingly. Remember to change the other parts of the script file too.
    ws['D{}'.format(i)] = str(v[3])
    ws['E{}'.format(i)] = str(v[4])
    ws['F{}'.format(i)] = str(v[5])
    ws['G{}'.format(i)] = str(v[6])

    wb.save(excel_file)  # Saves the excel file once the info is put in.


def get_file_info(file, x):  # Outputs a dictionary containing the data for worm x. This is the part of the script that actually gets the information from the HDF file.
    """
    This function grabs the info needed for worm x and sticks it into a dictionary.
    file = the path for the HDF file we're data mining
    x = which worm index (row) we're looking at
    """

    f = h5py.File(file, 'r')

    if 'features_means' not in f:
        print('This HDF file has no features_means group in it.')
    else:
        dct = {}
        means = f['features_means']  # Now the means data set is stored as a variable. This is so that we can dive into it.

        dct['Worm_index'] = means[x][0]  # This part of the script goes into the features_means section of the HDF file, goes to column 0, and saves the information in it as 'Worm_index'.
        dct['n_frames'] = means[x][1]  # This goes into column 1 in the features_means (n_frames) and gets info. Remember that the first column is column 0, not column 1.
        dct['backward_time'] = means[x][723]  # If you count starting at 0 inside features_means, you'll find that backward_time is the 723rd column.
        dct['backward_distance'] = means[x][724]
        dct['paused_time'] = means[x][716]
        dct['forward_time'] = means[x][709]
        dct['forward_distance'] = means[x][710]  # To grab more information, you just need to add a new line below this that has the same formats as these lines.

        return dct


def get_worm_indexes():  # Returns a tuple with a list of worm indexes and the file location.
    print('Type "quit" to exit.')  # This loop lets you put in worm numbers and outputs the data at each location.
    while True:
        f = str(input('What is the path for the file? : '))
        if f == 'quit':
            break
        while True:
            print('For worm numbers 0, 1, and 2, worm numbers should be printed like this: 0 1 2')
            index_strings = (input('What are the worm numbers? (starts with the 0th worm at index 2) : '))
            index_list = index_strings.split(' ')
            xname = input('What should the excel file be called? (Files should be named like this: name.xlsx) : ')
            return f, index_list, xname  # Returns a tuple with worm indexes and the file location.


tup = get_worm_indexes()
fl = tup[0]  # File location
index_ls = tup[1]  # List containing the worm numbers we want to grab
name = tup[2]
make_workbook(name)

for i in range(len(index_ls)):  # Loops a number of times equal to the number of worms we want to grab. First 'i' is 0.
    current_info_dict = get_file_info(fl, int(index_ls[i]))  # Dictionary with the info we currently want.
    excel_row = i + 2  # Because we want to start adding on cell A2
    print_to_excel(name, current_info_dict, excel_row)  # Adds the info to excel!
