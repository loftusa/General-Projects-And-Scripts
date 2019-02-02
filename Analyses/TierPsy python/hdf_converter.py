import h5py
import openpyxl


def make_workbook(excel_name):  # Finish this in a bit.
    wb = openpyxl.Workbook()
    ws = wb.active

    ws['A1'] = 'Worm_Index'
    ws['B1'] = 'n_frames'
    ws['C1'] = 'Backwards Time'
    ws['D1'] = 'Backwards Distance'
    ws['E1'] = 'Paused Time'
    ws['F1'] = 'Forward Time'
    ws['G1'] = 'Forward Distance'

    wb.save(str(excel_name))


def print_to_excel(excel_file, dct, i):  # This will open up an excel file and print each value of a dictionary into it column-wise.
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    v = list(dct.values())

    ws['A{}'.format(i)] = str(v[0])
    ws['B{}'.format(i)] = str(v[1])
    ws['C{}'.format(i)] = str(v[2])
    ws['D{}'.format(i)] = str(v[3])
    ws['E{}'.format(i)] = str(v[4])
    ws['F{}'.format(i)] = str(v[5])
    ws['G{}'.format(i)] = str(v[6])

    wb.save(excel_file)


def get_file_info(file, x):  # Outputs a dictionary containing the data for worm x.
    """
    This function grabs the info needed for worm x and sticks it into a dictionary.
    file = the path for the HDF file we're data mining
    x = which worm index (row) we're looking at
    """

    f = h5py.File(file, 'r')

    if 'features_means' not in f:
        print('This HDF file has no features_means group in it.')
    else:
        dct = {}
        means = f['features_means']  # Now the means data set is stored as a variable

        dct['Worm_index'] = means[x][0]
        dct['n_frames'] = means[x][1]
        dct['backward_time'] = means[x][723]
        dct['backward_distance'] = means[x][724]
        dct['paused_time'] = means[x][716]
        dct['forward_time'] = means[x][709]
        dct['forward_distance'] = means[x][710]

        return dct


def get_worm_indexes():  # Returns a tuple with a list of worm indexes and the file location.
    print('Type "quit" to exit.')  # This loop lets you put in worm numbers and outputs the data at each location.
    f = str(input('What is the path for the file? : '))
    print('For worm numbers 0, 1, and 2, worm numbers should be printed like this: 0 1 2')
    index_strings = (input('What are the worm numbers? (starts with the 0th worm at index 2) : '))
    index_list = index_strings.split(' ')
    xname = input('What should the excel file be called? (Files should be named like this: name.xlsx) : ')
    return f, index_list, xname  # Returns a tuple with worm indexes and the file location.


def print_worm_data(file, indx):  # 'file' is the file location. 'data' is a list containing the worm numbers to get data from
    for i in range(len(indx)):
        print('\nWORM {}: '.format(i) + str(get_file_info(file, int(indx[i]))))


tup = get_worm_indexes()
fl = tup[0]  # File location
index_ls = tup[1]  # List containing the worm numbers we want to grab
name = tup[2]  # Name of the excel file
make_workbook(name)

for i in range(len(index_ls)):  # Loops a number of times equal to the number of worms we want to grab. First 'i' is 0.
    current_info_dict = get_file_info(fl, int(index_ls[i]))  # Dictionary with the info we currently want.
    excel_row = i + 2  # Because we want to start adding on cell A2
    print_to_excel(name, current_info_dict, excel_row)  # Adds the info to excel!
