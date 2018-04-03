import h5py
import openpyxl
import os


def make_workbook(excel_name):  # This function makes a new workbook and puts the below names into their respective cells
    wb = openpyxl.Workbook()
    ws = wb.active

    ws['A1'] = 'Worm_Index'  # puts 'Worm_Index' on to cell A1. The below lines work equivalently.
    ws['B1'] = 'n_frames'
    ws['C1'] = 'Backwards Time'
    ws['D1'] = 'Backwards Distance'
    ws['E1'] = 'Paused Time'
    ws['F1'] = 'Forward Time'
    ws['G1'] = 'Forward Distance'
    ws['H1'] = 'Length'
    ws['I1'] = 'Area'
    ws['J1'] = 'Num_Backward_Events'
    ws['K1'] = 'Num_Forward_Events'
    ws['L1'] = 'Num_Omega_Events'
    ws['M1'] = 'Num_Paused_Events'
    ws['O1'] = 'Worm_Index'
    ws['P1'] = 'Timestamp'
    ws['Q1'] = 'Motion_Modes'

    wb.save(str(excel_name))  # Saves the excel file once the column headers have been added.


def print_to_excel(excel_file, dct, i):  # This function will open up an excel file and print each value of dct (the actual info) into it column-wise.
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    v = list(dct.values())

    ws['A{}'.format(i)] = str(v[0])  # This is the part that puts the info for the current worm into the current column. 'A{}' is A2, A3, A4, etc depending on which worm's info is currently being added.
    ws['B{}'.format(i)] = str(v[1])  # Works equivalently for the above, but with cell B1, B2, etc.
    ws['C{}'.format(i)] = str(v[2])  # If you want to make this script fill out a new cell, copy/paste one of these lines and adjust accordingly. Remember to change the other parts of the script file too.
    ws['D{}'.format(i)] = str(v[3])
    ws['E{}'.format(i)] = str(v[4])
    ws['F{}'.format(i)] = str(v[5])
    ws['G{}'.format(i)] = str(v[6])
    ws['H{}'.format(i)] = str(v[7])
    ws['I{}'.format(i)] = str(v[8])

    wb.save(excel_file)  # Saves the excel file once the info is put in.


def print_to_excel_timeseries(excel_file, tim, i):  # This function will open up an excel file and print each value of dct (the actual info) into it column-wise.
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    z = list(tim.values())
    
    ws['O{}'.format(i)] = str(z[0])
    ws['P{}'.format(i)] = str(z[1])
    ws['Q{}'.format(i)] = str(z[2])

    wb.save(excel_file)  # Saves the excel file once the info is put in.


def get_file_info(file, x):  # Outputs a dictionary containing the data for worm x. This is the part of the script that actually gets the information from the HDF file.
    """
    This function grabs the info needed for worm x and sticks it into a dictionary.
    file = the path for the HDF file we're data mining
    x = which worm index (row) we're looking at
    """

    f = h5py.File(file, 'r')

    if 'features_means' not in f:
        print('This HDF file has no features_means group in it.')
    else:
        dct = {}
        means = f['features_means']  # Now the means data set is stored as a variable. This is so that we can dive into it.

        dct['Worm_index'] = means[x][0]  # This part of the script goes into the features_means section of the HDF file, goes to column 0, and saves the information in it as 'Worm_index'.
        dct['n_frames'] = means[x][1]  # This goes into column 1 in the features_means (n_frames) and gets info. Remember that the first column is column 0, not column 1.
        dct['backward_time'] = means[x][723]  # If you count starting at 0 inside features_means, you'll find that backward_time is the 723rd column.
        dct['backward_distance'] = means[x][724]
        dct['paused_time'] = means[x][716]
        dct['forward_time'] = means[x][709]
        dct['forward_distance'] = means[x][710]  # To grab more information, you just need to add a new line below this that has the same formats as these lines.
        dct['length'] = means[x][8]
        dct['area'] = means[x][24]

        return dct


def get_file_info_timeseries(file, x):  # Outputs a dictionary containing the data for worm x. This is the part of the script that actually gets the information from the HDF file.
    """
    This function grabs the info needed for worm x and sticks it into a dictionary.
    file = the path for the HDF file we're data mining
    x = which worm index (row) we're looking at
    """

    f = h5py.File(file, 'r')

    if 'features_timeseries' not in f:
        print('This HDF file has no features_timeseries group in it.')
    else:
        tim = {}
        means = f['features_timeseries']  # Now the means data set is stored as a variable. This is so that we can dive into it.

        tim['Worm_index'] = means[x][0]  # This part of the script goes into the features_timeseries section of the HDF file, goes to column 0, and saves the information in it as 'Worm_index'.
        tim['Timestamp'] = means[x][1]
        tim['Motion_modes'] = means[x][3]
        
        return tim
    

def get_worm_indexes():  # Returns a tuple with a list of worm indexes and the file location.
    print('Type "quit" to exit.')  # This loop lets you put in worm numbers and outputs the data at each location.
    while True:
        f = str(input('What is the path for the file? : '))
        if f == 'quit':
            break
        while True:
            f2 = h5py.File(f, 'r')
            numMembers = f2['features_events']
            lNm = len(numMembers)
            rL = range(0, int(lNm))
            index_strings = ' '.join(str(x) for x in rL)
            index_list = index_strings.split(' ')
            print("Rows in features_means table: ", index_strings)
            rowNum = f2['features_timeseries']
            a = range(0, int(len(rowNum) - 1))
            print("Number of rows in timeseries table: ", int(len(rowNum)), "(rows: 0 through ", int(len(rowNum) - 1), ")")
            timeseries_end_row = ' '.join(str(x) for x in a)
            timeseries_list = timeseries_end_row.split(' ')
            xname = str(str(os.path.basename(f)) + '.xlsx')
            print("saving... console will close when complete")
            return f, index_list, timeseries_list, xname  # Returns a tuple with worm indexes and the file location.


tup = get_worm_indexes()
fl = tup[0]  # File location
index_ls = tup[1]  # List containing the worm numbers we want to grab
timeseries_ls = tup[2]
name = tup[3]
make_workbook(name)

for i in range(len(index_ls)):  # Loops a number of times equal to the number of worms we want to grab. First 'i' is 0.
    current_info_dict = get_file_info(fl, int(index_ls[i]))  # Dictionary with the info we currently want.
    excel_row = i + 2  # Because we want to start adding on cell A2
    print_to_excel(name, current_info_dict, excel_row)  # Adds the info to excel!

for i in range(len(timeseries_ls)):  # Loops a number of times equal to the amount of rows in features_timeseries.
    current_info_dict_timeseries = get_file_info_timeseries(fl, int(timeseries_ls[i]))  # Dictionary with the info we currently want.
    excel_row = i + 2  # Because we want to start adding on cell A2
    print_to_excel_timeseries(name, current_info_dict_timeseries, excel_row)  # Adds the info to excel!    
